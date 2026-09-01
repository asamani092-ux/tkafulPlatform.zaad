"""Reporting API views (moved from volunteering — Phase 1 API ownership)."""
from decimal import Decimal
from io import BytesIO

import openpyxl
from django.contrib.auth.models import User
from django.db.models import Sum, Q
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from core.permissions import IsAdmin
from rest_framework.permissions import AllowAny

from volunteering.models import VolunteeringProfile, ProjectAssignment, Task
from volunteering.serializers import TaskSerializer

from .models import (
    AdminReport,
    VolunteerStatistics,
    QuarterlyTarget,
    DepartmentHours,
    TopVolunteer,
)
from .serializers import AdminReportSerializer, VolunteerStatisticsSerializer


@api_view(['GET'])
@permission_classes([IsAdmin])
def projects_progress_report(request):
    """
    GET /api/admin/reports/projects-progress/
    Returns progress report for all projects
    """
    profiles = VolunteeringProfile.objects.filter(
        volunteer_status__in=['ACTIVE', 'COMPLETED']
    ).select_related("project").values('project__name', 'progress')

    return Response({
        'projects': [
            {
                'name': p['project__name'],
                'progress': p['progress']
            }
            for p in profiles
        ]
    })


@api_view(['GET'])
@permission_classes([IsAdmin])
def volunteers_performance_report(request):
    """
    GET /api/admin/reports/volunteers-performance/
    Returns performance report for all volunteers
    """
    volunteers = User.objects.filter(
        profile__role='user',
        profile__is_approved=True
    ).select_related('profile')

    data = []
    for volunteer in volunteers:
        completed = volunteer.assigned_tasks.filter(status='مكتملة').count()
        current = volunteer.assigned_tasks.exclude(status='مكتملة').count()
        total = completed + current
        completion_rate = int((completed / total) * 100) if total > 0 else 0

        data.append({
            'name': volunteer.profile.name,
            'completed': completed,
            'current': current,
            'completion_rate': completion_rate,
            'join_date': volunteer.date_joined.strftime('%Y-%m-%d'),
        })

    return Response({
        'volunteers': data
    })


@api_view(['GET'])
@permission_classes([IsAdmin])
def volunteer_tasks_report(request):
    """
    GET /api/admin/reports/volunteer-tasks/?volunteer_id=1
    Returns tasks for a specific volunteer
    """
    volunteer_id = request.query_params.get('volunteer_id')

    if not volunteer_id:
        return Response(
            {"error": "volunteer_id parameter is required"},
            status=status.HTTP_400_BAD_REQUEST
        )

    tasks = Task.objects.filter(
        volunteer_id=volunteer_id
    ).select_related('project').order_by('-due_date')

    serializer = TaskSerializer(tasks, many=True)

    # مقاييس الإنجاز المطلوبة (UX2 P4 · المتطوّعون 3.4):
    # ساعات التطوّع + عدد المشاريع المشارَك فيها + المهام المنجزة.
    completed_tasks = sum(1 for t in tasks if t.status == 'مكتملة')
    projects_participated = len({t.project_id for t in tasks if t.project_id})
    profile = getattr(User.objects.filter(pk=volunteer_id).first(), 'profile', None)
    volunteer_hours = getattr(profile, 'total_volunteer_hours', 0) if profile else 0

    return Response({
        'tasks': serializer.data,
        'metrics': {
            'volunteer_hours': volunteer_hours,
            'projects_participated': projects_participated,
            'completed_tasks': completed_tasks,
            'total_tasks': tasks.count(),
        },
    })


@api_view(['POST'])
@permission_classes([IsAdmin])
def generate_report(request):
    """
    POST /api/admin/reports/generate/
    Generate a comprehensive platform report
    Body: { "date_from": "2026-01-01", "date_to": "2026-01-31" } (optional)
    """
    date_from = request.data.get('date_from')
    date_to = request.data.get('date_to')

    project_filter = Q()
    task_filter = Q()
    volunteer_filter = Q()

    if date_from:
        project_filter &= Q(created_at__gte=date_from)
        task_filter &= Q(created_at__gte=date_from)
        volunteer_filter &= Q(date_joined__gte=date_from)

    if date_to:
        project_filter &= Q(created_at__lte=date_to)
        task_filter &= Q(created_at__lte=date_to)
        volunteer_filter &= Q(date_joined__lte=date_to)

    profiles = VolunteeringProfile.objects.filter(project_filter).select_related("project")
    projects_by_status = {
        'active': profiles.filter(volunteer_status='ACTIVE').count(),
        'completed': profiles.filter(volunteer_status='COMPLETED').count(),
        'planned': profiles.filter(volunteer_status='PLANNED').count(),
        'cancelled': profiles.filter(volunteer_status='CANCELLED').count(),
    }
    projects_by_category = {}
    for category in profiles.values_list('category', flat=True).distinct():
        if category:
            projects_by_category[category] = profiles.filter(category=category).count()

    projects_list = []
    for profile in profiles:
        platform = profile.project
        volunteers_count = ProjectAssignment.objects.filter(project=platform).count()
        tasks_count = Task.objects.filter(project=platform).count()
        tasks_completed = Task.objects.filter(project=platform, status='مكتملة').count()

        if tasks_count > 0:
            automatic_progress = int((tasks_completed / tasks_count) * 100)
        else:
            automatic_progress = profile.progress

        projects_list.append({
            'id': platform.id,
            'title': platform.name,
            'category': profile.category,
            'status': profile.volunteer_status,
            'status_display': {
                'ACTIVE': 'نشط',
                'COMPLETED': 'مكتمل',
                'PLANNED': 'متوقف',
                'CANCELLED': 'ملغي'
            }.get(profile.volunteer_status, profile.volunteer_status),
            'progress': automatic_progress,
            'beneficiaries': profile.beneficiaries,
            'donation_amount': float(profile.donation_amount),
            'start_date': platform.start_date.isoformat() if platform.start_date else None,
            'end_date': platform.end_date.isoformat() if platform.end_date else None,
            'volunteers_assigned': volunteers_count,
            'tasks_total': tasks_count,
            'tasks_completed': tasks_completed,
            'completion_rate': automatic_progress,
        })

    volunteers = User.objects.filter(
        profile__role='user',
        profile__is_approved=True
    ).filter(volunteer_filter).select_related('profile')

    volunteers_list = []
    for volunteer in volunteers:
        tasks = volunteer.assigned_tasks.all()
        tasks_completed = tasks.filter(status='مكتملة').count()
        tasks_in_progress = tasks.exclude(status='مكتملة').count()
        current_projects = list(set([task.project.name for task in tasks.exclude(status='مكتملة')]))

        volunteers_list.append({
            'id': volunteer.id,
            'name': volunteer.profile.name,
            'email': volunteer.email,
            'phone': volunteer.profile.phone,
            'city': volunteer.profile.city,
            'skills': volunteer.profile.skills,
            'qualification': volunteer.profile.qualification,
            'university': volunteer.profile.university,
            'total_hours': volunteer.profile.total_volunteer_hours,
            'tasks_completed': tasks_completed,
            'tasks_in_progress': tasks_in_progress,
            'rating': float(volunteer.profile.rating),
            'join_date': volunteer.date_joined.isoformat(),
            'current_projects': current_projects,
        })

    tasks = Task.objects.filter(task_filter).select_related('project', 'volunteer')
    tasks_by_status = {
        'in_progress': tasks.filter(status='قيد التنفيذ').count(),
        'waiting': tasks.filter(status='في الانتظار').count(),
        'completed': tasks.filter(status='مكتملة').count(),
        'on_hold': tasks.filter(status='معلقة').count(),
    }
    tasks_by_priority = {
        'high': tasks.filter(priority='عالية').count(),
        'medium': tasks.filter(priority='متوسطة').count(),
        'low': tasks.filter(priority='منخفضة').count(),
    }

    overdue_tasks = tasks.filter(
        due_date__lt=timezone.now().date(),
        status__in=['قيد التنفيذ', 'في الانتظار', 'معلقة']
    )
    overdue_tasks_list = [{
        'id': task.id,
        'title': task.title,
        'project': task.project.name,
        'volunteer': task.volunteer.profile.name if task.volunteer else None,
        'due_date': task.due_date.isoformat() if task.due_date else None,
        'status': task.status,
        'priority': task.priority,
    } for task in overdue_tasks[:10]]

    total_beneficiaries = profiles.aggregate(total=Sum('beneficiaries'))['total'] or 0
    total_donations = profiles.aggregate(total=Sum('donation_amount'))['total'] or 0
    total_volunteer_hours = volunteers.aggregate(
        total=Sum('profile__total_volunteer_hours')
    )['total'] or 0

    report_data = {
        'summary': {
            'total_projects': profiles.count(),
            'total_volunteers': volunteers.count(),
            'total_tasks': tasks.count(),
            'total_beneficiaries': total_beneficiaries,
            'total_donations': float(total_donations),
            'total_volunteer_hours': total_volunteer_hours,
            'date_from': date_from,
            'date_to': date_to,
        },
        'projects': {
            'by_status': projects_by_status,
            'by_category': projects_by_category,
            'list': projects_list,
        },
        'volunteers': {
            'total': volunteers.count(),
            'list': volunteers_list,
        },
        'tasks': {
            'by_status': tasks_by_status,
            'by_priority': tasks_by_priority,
            'overdue': overdue_tasks_list,
            'total_completed': tasks_by_status['completed'],
            'completion_rate': int((tasks_by_status['completed'] / tasks.count()) * 100) if tasks.count() > 0 else 0,
        },
    }

    now = timezone.now()
    if date_from and date_to:
        title = f"تقرير شامل ({date_from} - {date_to})"
    else:
        title = f"تقرير شامل - {now.strftime('%Y-%m-%d %H:%M')}"

    report = AdminReport.objects.create(
        admin=request.user,
        title=title,
        date_from=date_from,
        date_to=date_to,
        report_data=report_data,
        total_projects=profiles.count(),
        total_volunteers=volunteers.count(),
        total_tasks=tasks.count(),
        total_beneficiaries=total_beneficiaries,
        total_donations=total_donations,
    )

    serializer = AdminReportSerializer(report)
    return Response({
        'message': 'تم إنشاء التقرير بنجاح',
        'report': serializer.data
    }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAdmin])
def list_reports(request):
    """
    GET /api/admin/reports/
    List all generated reports
    """
    reports = AdminReport.objects.all()
    serializer = AdminReportSerializer(reports, many=True)
    return Response({'results': serializer.data})


@api_view(['GET'])
@permission_classes([IsAdmin])
def get_report_detail(request, report_id):
    """
    GET /api/admin/reports/{id}/
    Get specific report details
    """
    try:
        report = AdminReport.objects.get(id=report_id)
        serializer = AdminReportSerializer(report)
        return Response(serializer.data)
    except AdminReport.DoesNotExist:
        return Response(
            {'error': 'التقرير غير موجود'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['DELETE'])
@permission_classes([IsAdmin])
def delete_report(request, report_id):
    """
    DELETE /api/admin/reports/{id}/
    Delete a report
    """
    try:
        report = AdminReport.objects.get(id=report_id)
        report_title = report.title
        report.delete()
        return Response({
            'message': f'تم حذف التقرير "{report_title}" بنجاح'
        })
    except AdminReport.DoesNotExist:
        return Response(
            {'error': 'التقرير غير موجود'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['GET'])
@permission_classes([IsAdmin])
def report_scope(request):
    """
    GET /api/reports/scope/?type=<platform|project|volunteers|sponsorships>&project=<id|slug>
    بوّابة تقارير موحّدة (UX2 P4 · 3.9) — تُعيد بيانات منظّمة للعرض على الشاشة
    وتصدير CSV/PDF على الواجهة (المتصفّح يشكّل العربية طباعياً بشكل صحيح).
    O(N) على صفوف النطاق المطلوب.
    """
    from projects.models import Project

    scope = request.query_params.get('type', 'platform')
    project_ref = request.query_params.get('project')

    def project_row(p):
        tasks = Task.objects.filter(project=p)
        total = tasks.count()
        done = tasks.filter(status='مكتملة').count()
        return {
            'id': p.id,
            'name': p.name,
            'slug': p.slug,
            'status': p.status,
            'volunteers': ProjectAssignment.objects.filter(project=p).count(),
            'tasks_total': total,
            'tasks_completed': done,
            'completion_rate': int((done / total) * 100) if total else 0,
        }

    if scope == 'project':
        if not project_ref:
            return Response({'error': 'معرّف المشروع مطلوب'}, status=status.HTTP_400_BAD_REQUEST)
        p = Project.objects.filter(slug=project_ref).first() or Project.objects.filter(pk=project_ref if str(project_ref).isdigit() else 0).first()
        if not p:
            return Response({'error': 'المشروع غير موجود'}, status=status.HTTP_404_NOT_FOUND)
        rows = [project_row(p)]
        return Response({'scope': 'project', 'title': f'تقرير المشروع: {p.name}',
                         'columns': ['name', 'status', 'volunteers', 'tasks_total', 'tasks_completed', 'completion_rate'],
                         'rows': rows})

    if scope == 'volunteers':
        vols = User.objects.filter(profile__role='user', profile__is_approved=True).select_related('profile')
        rows = []
        for v in vols:
            completed = v.assigned_tasks.filter(status='مكتملة').count()
            projects_participated = len({t.project_id for t in v.assigned_tasks.all() if t.project_id})
            rows.append({
                'id': v.id,
                'name': v.profile.name,
                'email': v.email,
                'volunteer_hours': v.profile.total_volunteer_hours,
                'projects_participated': projects_participated,
                'completed_tasks': completed,
            })
        return Response({'scope': 'volunteers', 'title': 'تقرير المتطوّعين',
                         'columns': ['name', 'email', 'volunteer_hours', 'projects_participated', 'completed_tasks'],
                         'rows': rows})

    if scope == 'sponsorships':
        try:
            from sponsorships.models import Sponsorship
        except Exception:
            return Response({'scope': 'sponsorships', 'title': 'تقرير الكفالات', 'columns': [], 'rows': []})
        qs = Sponsorship.objects.select_related('donor__profile', 'project')
        if project_ref:
            qs = qs.filter(project__slug=project_ref)
        rows = [{
            'id': s.id,
            'type': s.type,
            'amount': float(s.amount),
            'status': s.status,
            'donor': getattr(getattr(s.donor, 'profile', None), 'name', '') or (s.donor.email if s.donor_id else ''),
            'project': s.project.name if s.project_id else '—',
        } for s in qs]
        return Response({'scope': 'sponsorships', 'title': 'تقرير الكفالات',
                         'columns': ['type', 'amount', 'status', 'donor', 'project'],
                         'rows': rows})

    # platform (افتراضي)
    projects = Project.objects.all()
    rows = [project_row(p) for p in projects]
    summary = {
        'total_projects': projects.count(),
        'total_volunteers': User.objects.filter(profile__role='user', profile__is_approved=True).count(),
        'total_tasks': Task.objects.count(),
        'total_completed_tasks': Task.objects.filter(status='مكتملة').count(),
        'total_hours': User.objects.filter(profile__role='user').aggregate(t=Sum('profile__total_volunteer_hours'))['t'] or 0,
    }
    return Response({'scope': 'platform', 'title': 'تقرير المنصّة الشامل', 'summary': summary,
                     'columns': ['name', 'status', 'volunteers', 'tasks_total', 'tasks_completed', 'completion_rate'],
                     'rows': rows})


@api_view(['GET'])
@permission_classes([AllowAny])
def public_volunteer_statistics(request):
    """
    GET /api/public-volunteer-statistics/
    Returns volunteer statistics for the home page dashboard
    Optional query param: ?year=2025 (defaults to latest year)
    """
    year = request.query_params.get('year', None)

    if year:
        stats = VolunteerStatistics.objects.filter(year=int(year)).first()
    else:
        stats = VolunteerStatistics.objects.first()

    if not stats:
        return Response({
            'error': 'No statistics found',
            'data': None
        }, status=status.HTTP_404_NOT_FOUND)

    serializer = VolunteerStatisticsSerializer(stats)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAdmin])
def upload_volunteer_statistics(request):
    """
    POST /api/admin/upload-statistics/
    Upload Excel file to update volunteer statistics for home page
    Requires admin authentication
    """
    if 'file' not in request.FILES:
        return Response({
            'success': False,
            'error': 'لم يتم رفع أي ملف'
        }, status=status.HTTP_400_BAD_REQUEST)

    excel_file = request.FILES['file']

    try:
        wb = openpyxl.load_workbook(BytesIO(excel_file.read()), data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        total_records = len([r for r in rows if r and any(r)])

        volunteer_ids = set()
        total_hours = 0
        department_hours_map = {}
        volunteer_hours_map = {}

        headers = [cell.value for cell in ws[1]]

        hours_col = None
        dept_col = None
        name_col = None
        id_col = None

        for i, h in enumerate(headers):
            if h:
                h_lower = str(h).strip()
                if 'ساع' in h_lower or 'hours' in h_lower.lower():
                    hours_col = i
                elif 'إدار' in h_lower or 'قسم' in h_lower or 'department' in h_lower.lower():
                    dept_col = i
                elif 'اسم' in h_lower and 'مشروع' not in h_lower:
                    name_col = i
                elif 'هوية' in h_lower or 'id' in h_lower.lower():
                    id_col = i

        for row in rows:
            if not row or not any(row):
                continue

            vol_id = None
            if id_col is not None and id_col < len(row):
                vol_id = row[id_col]
            if vol_id:
                volunteer_ids.add(str(vol_id))

            hours = 0
            if hours_col is not None and hours_col < len(row):
                try:
                    hours = float(row[hours_col] or 0)
                except (ValueError, TypeError):
                    hours = 0
            total_hours += hours

            dept = 'غير مسند'
            if dept_col is not None and dept_col < len(row):
                dept = str(row[dept_col] or 'غير مسند').strip()

            if dept not in department_hours_map:
                department_hours_map[dept] = 0
            department_hours_map[dept] += hours

            vol_name = None
            if name_col is not None and name_col < len(row):
                vol_name = str(row[name_col] or '').strip()
            if vol_name:
                if vol_name not in volunteer_hours_map:
                    volunteer_hours_map[vol_name] = 0
                volunteer_hours_map[vol_name] += hours

        year = 2025
        stats, created = VolunteerStatistics.objects.update_or_create(
            year=year,
            defaults={
                'total_volunteers': len(volunteer_ids) or total_records,
                'new_volunteers': int(len(volunteer_ids) * 0.78),
                'returning_volunteers': int(len(volunteer_ids) * 0.22),
                'total_hours': int(total_hours),
                'total_contribution_value': Decimal(total_hours * 13),
                'contribution_value_display': (
                    f"{total_hours * 13 / 1000000:.2f}M"
                    if total_hours > 100000
                    else f"{total_hours * 13 / 1000:.0f}K"
                ),
            }
        )

        DepartmentHours.objects.filter(statistics=stats).delete()

        colors = ['#6B1F2B', '#8B5A2B', '#2E8B57', '#4169E1', '#9370DB', '#FF8C00', '#20B2AA', '#DC143C']
        total_dept_hours = sum(department_hours_map.values()) or 1

        for i, (dept, hours) in enumerate(sorted(department_hours_map.items(), key=lambda x: -x[1])):
            DepartmentHours.objects.create(
                statistics=stats,
                department_name=dept,
                department_name_ar=dept,
                hours=int(hours),
                percentage=Decimal(hours / total_dept_hours * 100),
                color=colors[i % len(colors)]
            )

        TopVolunteer.objects.filter(statistics=stats).delete()

        top_vols = sorted(volunteer_hours_map.items(), key=lambda x: -x[1])[:5]
        for rank, (name, hours) in enumerate(top_vols, 1):
            TopVolunteer.objects.create(
                statistics=stats,
                rank=rank,
                name=name,
                hours=int(hours)
            )

        return Response({
            'success': True,
            'message': 'تم تحديث الإحصائيات بنجاح',
            'data': {
                'total_records': total_records,
                'total_volunteers': len(volunteer_ids) or total_records,
                'total_hours': int(total_hours),
                'departments': len(department_hours_map),
                'top_volunteers': len(top_vols)
            }
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({
            'success': False,
            'error': f'خطأ في معالجة الملف: {str(e)}'
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT'])
@permission_classes([IsAdmin])
def admin_volunteer_statistics(request):
    """
    GET/PUT /api/admin/volunteer-statistics/
    Get or update volunteer statistics for home page dashboard
    Requires admin authentication
    """
    year = request.query_params.get('year', 2025)

    if request.method == 'GET':
        stats = VolunteerStatistics.objects.filter(year=year).first()
        if not stats:
            return Response({'error': 'No statistics found'}, status=status.HTTP_404_NOT_FOUND)
        serializer = VolunteerStatisticsSerializer(stats)
        return Response(serializer.data)

    elif request.method == 'PUT':
        data = request.data

        stats, created = VolunteerStatistics.objects.update_or_create(
            year=int(data.get('year', 2025)),
            defaults={
                'total_volunteers': int(data.get('total_volunteers', 0)),
                'new_volunteers': int(data.get('new_volunteers', 0)),
                'returning_volunteers': int(data.get('returning_volunteers', 0)),
                'total_hours': int(data.get('total_hours', 0)),
                'total_contribution_value': Decimal(str(data.get('total_contribution_value', 0))),
                'contribution_value_display': data.get('contribution_value_display', ''),
            }
        )

        quarterly_targets = data.get('quarterly_targets', [])
        if quarterly_targets:
            QuarterlyTarget.objects.filter(statistics=stats).delete()
            for qt in quarterly_targets:
                QuarterlyTarget.objects.create(
                    statistics=stats,
                    quarter=int(qt.get('quarter', 1)),
                    volunteer_target=int(qt.get('volunteer_target', 0)),
                    volunteer_actual=int(qt.get('volunteer_actual', 0)),
                    hours_target=int(qt.get('hours_target', 0)),
                    hours_actual=int(qt.get('hours_actual', 0)),
                )

        department_hours = data.get('department_hours', [])
        if department_hours:
            DepartmentHours.objects.filter(statistics=stats).delete()
            for dh in department_hours:
                DepartmentHours.objects.create(
                    statistics=stats,
                    department_name=dh.get('department_name', ''),
                    department_name_ar=dh.get('department_name_ar', dh.get('label', '')),
                    hours=int(dh.get('hours', dh.get('value', 0))),
                    percentage=Decimal(str(dh.get('percentage', 0))),
                    color=dh.get('color', '#6B1F2B'),
                )

        top_volunteers = data.get('top_volunteers', [])
        if top_volunteers:
            TopVolunteer.objects.filter(statistics=stats).delete()
            for tv in top_volunteers:
                TopVolunteer.objects.create(
                    statistics=stats,
                    rank=int(tv.get('rank', 1)),
                    name=tv.get('name', ''),
                    hours=int(tv.get('hours', 0)),
                )

        serializer = VolunteerStatisticsSerializer(stats)
        return Response({
            'success': True,
            'message': 'تم تحديث الإحصائيات بنجاح',
            'data': serializer.data
        }, status=status.HTTP_200_OK)
