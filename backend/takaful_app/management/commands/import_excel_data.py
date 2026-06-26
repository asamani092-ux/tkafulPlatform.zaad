"""
Management command to import data from Excel files.
Imports projects and volunteers from the provided Excel files.
"""
import os
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from django.db import transaction
from accounts.models import Profile
from takaful_app.models import Project
import openpyxl
from datetime import datetime


class Command(BaseCommand):
    help = 'Import projects and volunteers from Excel files'

    def add_arguments(self, parser):
        parser.add_argument(
            '--projects-only',
            action='store_true',
            help='Import only projects',
        )
        parser.add_argument(
            '--volunteers-only',
            action='store_true',
            help='Import only volunteers',
        )

    def handle(self, *args, **options):
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

        takaful_file = os.path.join(base_dir, 'data_takaful.xlsx')

        if not os.path.exists(takaful_file):
            self.stderr.write(self.style.ERROR(f'File not found: {takaful_file}'))
            return

        projects_only = options.get('projects_only', False)
        volunteers_only = options.get('volunteers_only', False)

        # Load workbook
        self.stdout.write(f'Loading workbook: {takaful_file}')
        wb = openpyxl.load_workbook(takaful_file, data_only=True)

        self.stdout.write(f'Available sheets: {wb.sheetnames}')

        if not volunteers_only:
            self.import_projects(wb)

        if not projects_only:
            self.import_volunteers(wb)

        self.stdout.write(self.style.SUCCESS('Import completed successfully!'))

    def parse_date(self, value):
        """Parse date from various formats"""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            # Try different date formats
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return None

    def parse_decimal(self, value):
        """Parse decimal from various formats"""
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, str):
            value = value.strip().replace(',', '').replace('ريال', '').strip()
            try:
                return float(value)
            except ValueError:
                return 0
        return 0

    def get_category(self, project_type):
        """Map project type to category"""
        if not project_type:
            return 'أساسي'
        project_type = str(project_type).strip()
        if 'مجتمعي' in project_type:
            return 'مجتمعي'
        elif 'مؤسسي' in project_type:
            return 'مؤسسي'
        else:
            return 'أساسي'

    @transaction.atomic
    def import_projects(self, wb):
        """Import projects from المشروع sheet"""
        sheet_name = 'المشروع'
        if sheet_name not in wb.sheetnames:
            self.stderr.write(self.style.WARNING(f'Sheet "{sheet_name}" not found'))
            return

        ws = wb[sheet_name]

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        self.stdout.write(f'Project headers: {headers}')

        # Column mapping (Arabic to field)
        column_map = {
            'اسم المشروع': 'title',
            'نوع المشروع': 'category',
            'وصف المشروع': 'desc',
            'مبلغ التبرع للمشروع': 'donation_amount',
            'موقع تنفيذ المشروع': 'location',
            'تاريخ بداية المشروع': 'start_date',
            'تاريخ الانتهاء المتوقع': 'end_date',
            'متطلبات التنفيذ': 'implementation_requirements',
            'أهداف المشروع': 'project_goals',
        }

        # Find column indices
        col_indices = {}
        for i, header in enumerate(headers):
            if header in column_map:
                col_indices[column_map[header]] = i

        self.stdout.write(f'Column indices: {col_indices}')

        created_count = 0
        updated_count = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or not row[0]:
                continue

            title = str(row[col_indices.get('title', 0)] or '').strip()
            if not title:
                continue

            project_type = row[col_indices.get('category', 1)] if 'category' in col_indices else None

            project_data = {
                'desc': str(row[col_indices.get('desc', 2)] or '').strip(),
                'category': self.get_category(project_type),
                'donation_amount': self.parse_decimal(row[col_indices.get('donation_amount', 3)] if 'donation_amount' in col_indices else 0),
                'location': str(row[col_indices.get('location', 4)] or '').strip() if 'location' in col_indices else '',
                'start_date': self.parse_date(row[col_indices.get('start_date', 5)] if 'start_date' in col_indices else None),
                'end_date': self.parse_date(row[col_indices.get('end_date', 6)] if 'end_date' in col_indices else None),
                'implementation_requirements': str(row[col_indices.get('implementation_requirements', 7)] or '').strip() if 'implementation_requirements' in col_indices else '',
                'project_goals': str(row[col_indices.get('project_goals', 8)] or '').strip() if 'project_goals' in col_indices else '',
                'status': 'ACTIVE',
            }

            # Check if project exists
            project, created = Project.objects.update_or_create(
                title=title,
                defaults=project_data
            )

            if created:
                created_count += 1
                self.stdout.write(f'  Created project: {title}')
            else:
                updated_count += 1
                self.stdout.write(f'  Updated project: {title}')

        self.stdout.write(self.style.SUCCESS(
            f'Projects: {created_count} created, {updated_count} updated'
        ))

    @transaction.atomic
    def import_volunteers(self, wb):
        """Import volunteers from المتطوع sheet"""
        sheet_name = 'المتطوع'
        if sheet_name not in wb.sheetnames:
            self.stderr.write(self.style.WARNING(f'Sheet "{sheet_name}" not found'))
            return

        ws = wb[sheet_name]

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        self.stdout.write(f'Volunteer headers: {headers}')

        # Column mapping
        column_map = {
            'الاسم ثلاثي': 'name',
            'رقم الهوية': 'id_number',
            'البريد الالكتروني': 'email',
            'رقم الجوال': 'phone',
            'العمر': 'age',
            'الجنس': 'gender',
            'المنطقة': 'region',
            'المدينة': 'city',
            'المستوى التعليمي': 'qualification',
            'المهارات': 'skills',
        }

        # Find column indices
        col_indices = {}
        for i, header in enumerate(headers):
            if header in column_map:
                col_indices[column_map[header]] = i

        self.stdout.write(f'Column indices: {col_indices}')

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or not any(row):
                continue

            name = str(row[col_indices.get('name', 0)] or '').strip()
            email = str(row[col_indices.get('email', 2)] or '').strip().lower() if 'email' in col_indices else ''
            phone = str(row[col_indices.get('phone', 3)] or '').strip() if 'phone' in col_indices else ''

            if not name:
                skipped_count += 1
                continue

            # Generate email if missing
            if not email or email == 'none' or '@' not in email:
                # Create email from phone or name
                if phone:
                    clean_phone = ''.join(filter(str.isdigit, phone))
                    email = f"volunteer_{clean_phone}@takaful.local"
                else:
                    # Create from name
                    clean_name = name.replace(' ', '_').lower()
                    email = f"{clean_name}_{row_num}@takaful.local"

            # Parse age
            age_value = row[col_indices.get('age', 4)] if 'age' in col_indices else None
            age = None
            if age_value:
                try:
                    age = int(age_value)
                except (ValueError, TypeError):
                    pass

            # Parse gender
            gender = str(row[col_indices.get('gender', 5)] or '').strip() if 'gender' in col_indices else ''
            if gender and 'ذكر' in gender:
                gender = 'ذكر'
            elif gender and ('أنثى' in gender or 'انثى' in gender):
                gender = 'أنثى'
            else:
                gender = ''

            # Parse skills
            skills_raw = str(row[col_indices.get('skills', 9)] or '').strip() if 'skills' in col_indices else ''
            skills = [s.strip() for s in skills_raw.split(',') if s.strip()] if skills_raw else []

            city = str(row[col_indices.get('city', 7)] or '').strip() if 'city' in col_indices else ''
            qualification = str(row[col_indices.get('qualification', 8)] or '').strip() if 'qualification' in col_indices else ''

            try:
                # Check if user exists
                user, user_created = User.objects.get_or_create(
                    email=email,
                    defaults={
                        'username': email,
                        'first_name': name.split()[0] if name else '',
                        'last_name': ' '.join(name.split()[1:]) if name and len(name.split()) > 1 else '',
                        'is_active': True,
                    }
                )

                # Update or create profile
                profile, profile_created = Profile.objects.update_or_create(
                    user=user,
                    defaults={
                        'name': name,
                        'phone': phone,
                        'city': city,
                        'gender': gender,
                        'age': age,
                        'qualification': qualification,
                        'skills': skills,
                        'role': 'user',
                        'is_approved': True,  # Auto-approve imported volunteers
                    }
                )

                if user_created:
                    created_count += 1
                    if created_count <= 10:  # Only show first 10
                        self.stdout.write(f'  Created volunteer: {name} ({email})')
                else:
                    updated_count += 1

            except Exception as e:
                self.stderr.write(f'  Error importing row {row_num}: {e}')
                skipped_count += 1
                continue

        self.stdout.write(self.style.SUCCESS(
            f'Volunteers: {created_count} created, {updated_count} updated, {skipped_count} skipped'
        ))
