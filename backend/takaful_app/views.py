from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.views import APIView
from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from django.db.models import Sum, Count, Q
from django.utils import timezone

from .models import (
    Project, Service, ServiceRequest, ServiceVolunteerApplication, Volunteer, Suggestion,
    ProjectAssignment, Task, Subtask, AdminReport, VolunteerApplication,
    VolunteerStatistics, QuarterlyTarget, DepartmentHours, TopVolunteer, WaterSupplyRequest
)
import openpyxl
from io import BytesIO
from decimal import Decimal
from .serializers import (
    ProjectSerializer, ServiceSerializer, ServiceRequestSerializer, ServiceVolunteerApplicationSerializer,
    VolunteerSerializer, SuggestionSerializer, ProjectAssignmentSerializer,
    TaskSerializer, SubtaskSerializer, VolunteerDetailSerializer,
    VolunteerRequestSerializer, AdminReportSerializer, VolunteerApplicationSerializer,
    VolunteerStatisticsSerializer, WaterSupplyRequestSerializer
)


# Custom permission to check if user is admin
class IsAdmin(IsAuthenticated):
    def has_permission(self, request, view):
        return (
            super().has_permission(request, view) and
            request.user.profile.role == 'admin'
        )


# ============================================================================
# ADMIN STATISTICS ENDPOINTS
# ============================================================================
@api_view(['GET'])
@permission_classes([IsAdmin])
def admin_stats(request):
    """
    GET /api/admin/stats/
    Returns aggregated statistics for admin dashboard (main.tsx)
    """
    active_projects_count = Project.objects.filter(status='ACTIVE').count()
    completed_projects_count = Project.objects.filter(status='COMPLETED').count()
    total_projects_count = Project.objects.count()
    
    total_donations = Project.objects.aggregate(
        total=Sum('donation_amount')
    )['total'] or 0
    
    total_beneficiaries = Project.objects.aggregate(
        total=Sum('beneficiaries')
    )['total'] or 0
    
    return Response({
        'total_donations': float(total_donations),
        'total_beneficiaries': total_beneficiaries,
        'active_projects': active_projects_count,
        'completed_projects': completed_projects_count,
        'total_projects': total_projects_count,
    })


@api_view(['GET'])
@permission_classes([IsAdmin])
def volunteer_stats(request):
    """
    GET /api/admin/volunteer-stats/
    Returns volunteer statistics for VolunteerManagement page
    """
    # ✅ UPDATED: Count only APPROVED volunteers
    total_volunteers = User.objects.filter(
        profile__role='user',
        profile__is_approved=True
    ).count()
    
    # Active volunteers: those with current tasks
    active_volunteers = User.objects.filter(
        profile__role='user',
        profile__is_approved=True,
        assigned_tasks__status__in=['قيد التنفيذ', 'في الانتظار']
    ).distinct().count()
    
    # Total volunteer hours (only approved)
    total_hours = User.objects.filter(
        profile__role='user',
        profile__is_approved=True
    ).aggregate(
        total=Sum('profile__total_volunteer_hours')
    )['total'] or 0
    
    # Completed tasks
    completed_tasks = Task.objects.filter(status='مكتملة').count()
    
    return Response({
        'total_volunteers': total_volunteers,
        'active_volunteers': active_volunteers,
        'total_hours': total_hours,
        'completed_tasks': completed_tasks,
    })


@api_view(['GET'])
@permission_classes([IsAdmin])
def get_my_active_project(request):
    """
    GET /api/admin/my-active-project/?status=ACTIVE&project_id=123
    Returns a project for viewing/editing in the bottom section of main page

    Query Parameters:
    - status: Filter by status (ACTIVE, PLANNED, COMPLETED, CANCELLED)
              The dropdown menu uses this to SELECT which project to display
    - project_id: Return a specific project by ID

    Independent from top tabs section - used for filtering projects in bottom section
    """
    try:
        project_id = request.query_params.get('project_id')
        status_filter = request.query_params.get('status')

        # Status mapping from Arabic to English
        status_map = {
            'نشط': 'ACTIVE',
            'متوقف': 'PLANNED',
            'مكتمل': 'COMPLETED',
            'ملغي': 'CANCELLED'
        }

        # If specific project requested, return it (highest priority)
        if project_id:
            try:
                project = Project.objects.get(id=project_id)
                serializer = ProjectSerializer(project)
                return Response(serializer.data)
            except Project.DoesNotExist:
                pass

        # If status filter provided, return most recent project with that status
        if status_filter:
            # Convert Arabic status to English if needed
            english_status = status_map.get(status_filter, status_filter)

            project = Project.objects.filter(
                status=english_status
            ).order_by('-updated_at').first()

            if project:
                serializer = ProjectSerializer(project)
                return Response(serializer.data)
            else:
                # No project found with this status
                return Response(None, status=200)

        # Default: Return most recent active project, fallback to any recent project
        project = Project.objects.filter(status='ACTIVE').order_by('-updated_at').first()

        if not project:
            project = Project.objects.all().order_by('-updated_at').first()

        if project:
            serializer = ProjectSerializer(project)
            return Response(serializer.data)

        # Return null only if there are absolutely no projects
        return Response(None, status=200)

    except Exception as e:
        return Response({'error': str(e)}, status=500)


# ============================================================================
# PROJECT VIEWSET (Enhanced) - CORRECTED VERSION
# ============================================================================
class ProjectViewSet(viewsets.ModelViewSet):
    """
    Admin endpoint for managing projects
    Returns all projects (no pagination) filtered by status
    """
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [IsAdmin]
    pagination_class = None  # Disable pagination to return all projects

    STATUS_MAPPING = {
        'نشط': 'ACTIVE',
        'متوقف': 'PLANNED',
        'مكتمل': 'COMPLETED',
        'ملغي': 'CANCELLED',
        'حالة المشروع': 'PLANNED',
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        status_param = self.request.query_params.get('status', None)

        if status_param == 'pending':
            queryset = queryset.filter(status='PLANNED')
        elif status_param == 'active':
            queryset = queryset.filter(status='ACTIVE')
        elif status_param == 'completed':
            queryset = queryset.filter(status='COMPLETED')

        return queryset.order_by('-created_at')  # Most recent first

    def create(self, request, *args, **kwargs):
        """
        Create a new project - automatically set status to ACTIVE
        Projects created by admin should be active by default, not pending
        """
        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)

        # Set status to ACTIVE by default for admin-created projects
        if 'status' not in data:
            data['status'] = 'ACTIVE'

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        """Handle full updates (PUT)"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()

        # Make a mutable copy of the data
        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)

        # Handle status mapping from Arabic to English
        if 'status' in data:
            arabic_status = data['status']
            english_status = self.STATUS_MAPPING.get(arabic_status, arabic_status)
            data['status'] = english_status

        # Serialize and save
        serializer = self.get_serializer(instance, data=data, partial=True)

        # Add detailed error logging
        if not serializer.is_valid():
            print(f"❌ Validation Error for Project {instance.id}:")
            print(f"   Request data: {request.data}")
            print(f"   Processed data: {data}")
            print(f"   Errors: {serializer.errors}")

        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        # Return updated data (status_display is automatically added by serializer)
        return Response(serializer.data)
    
    def partial_update(self, request, *args, **kwargs):
        """Handle partial updates (PATCH)"""
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)
    
    @action(detail=True, methods=['post'])
    def assign_volunteer(self, request, pk=None):
        project = self.get_object()
        user_id = request.data.get('user_id')
        
        if not user_id:
            return Response(
                {"error": "user_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response(
                {"error": "User not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        assignment, created = ProjectAssignment.objects.get_or_create(
            project=project,
            user=user,
            defaults={'status': 'جديدة'}
        )
        
        if not created:
            return Response(
                {"error": "User already assigned to this project"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = ProjectAssignmentSerializer(assignment)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """
        Approve a pending project idea - changes status from PLANNED to ACTIVE
        POST /api/admin/projects/{id}/approve/
        """
        project = self.get_object()

        if project.status != 'PLANNED':
            return Response(
                {"error": "Only pending projects can be approved"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Change status from PLANNED (متوقف) to ACTIVE (نشط)
        old_status = project.status
        project.status = 'ACTIVE'
        project.save()

        print(f"✅ Project '{project.title}' approved: {old_status} → ACTIVE")

        serializer = self.get_serializer(project)
        return Response({
            "message": "Project approved successfully",
            "project": serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        project = self.get_object()
        
        if project.status != 'PLANNED':
            return Response(
                {"error": "Only pending projects can be rejected"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        project_title = project.title
        project.delete()
        
        return Response({
            "message": f"Project '{project_title}' rejected successfully"
        })


# ============================================================================
# TASK VIEWSET
# ============================================================================
class TaskViewSet(viewsets.ModelViewSet):
    """
    Admin endpoint for managing tasks
    GET /api/admin/tasks/ - List all tasks
    POST /api/admin/tasks/ - Create new task
    PUT /api/admin/tasks/{id}/ - Update task (including subtasks)
    """
    queryset = Task.objects.all()
    serializer_class = TaskSerializer
    permission_classes = [IsAdmin]
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('project', 'volunteer__profile')
        
        # Filter by volunteer
        volunteer_id = self.request.query_params.get('volunteer_id')
        if volunteer_id:
            queryset = queryset.filter(volunteer_id=volunteer_id)
        
        # Filter by project
        project_id = self.request.query_params.get('project_id')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        
        # Filter by status
        task_status = self.request.query_params.get('status')
        if task_status:
            queryset = queryset.filter(status=task_status)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def assign(self, request, pk=None):
        """
        POST /api/admin/tasks/{id}/assign/
        Body: { "volunteer_id": 1 }
        Assign a task to a volunteer
        """
        task = self.get_object()
        volunteer_id = request.data.get('volunteer_id')
        
        if not volunteer_id:
            return Response(
                {"error": "volunteer_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            volunteer = User.objects.get(id=volunteer_id)
        except User.DoesNotExist:
            return Response(
                {"error": "Volunteer not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        task.volunteer = volunteer
        if task.status == 'في الانتظار':
            task.status = 'قيد التنفيذ'
        task.save()
        
        serializer = self.get_serializer(task)
        return Response({
            "message": "Task assigned successfully",
            "task": serializer.data
        })


# ============================================================================
# VOLUNTEER MANAGEMENT ENDPOINTS
# ============================================================================
@api_view(['GET'])
@permission_classes([IsAdmin])
def list_volunteers(request):
    """
    GET /api/admin/volunteers/
    Returns detailed list of APPROVED volunteers for VolunteerManagement page
    """
    # ✅ UPDATED: Only show approved volunteers
    volunteers = User.objects.filter(
        profile__role='user',
        profile__is_approved=True
    ).select_related('profile').prefetch_related('assigned_tasks')
    
    serializer = VolunteerDetailSerializer(volunteers, many=True)
    return Response({
        'results': serializer.data
    })


@api_view(['GET'])
@permission_classes([IsAdmin])
def volunteer_requests_list(request):
    """
    GET /api/admin/volunteer-requests/?limit=4
    Returns list of PENDING volunteer approval requests
    For VolunteerRequests.tsx page and admin main page
    """
    # ✅ UPDATED: Only show pending (not approved) volunteers
    pending_volunteers = User.objects.filter(
        profile__role='user',
        profile__is_approved=False
    ).select_related('profile').order_by('-date_joined')

    # Support limit parameter
    limit = request.query_params.get('limit')
    if limit:
        try:
            pending_volunteers = pending_volunteers[:int(limit)]
        except ValueError:
            pass  # Ignore invalid limit values

    serializer = VolunteerRequestSerializer(pending_volunteers, many=True)
    return Response({
        'results': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsAdmin])
def accept_volunteer_request(request, volunteer_id):
    """
    POST /api/admin/volunteer-requests/{id}/accept/
    Accept a volunteer request
    """
    try:
        volunteer = User.objects.get(id=volunteer_id)
    except User.DoesNotExist:
        return Response(
            {"error": "Volunteer not found"},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # ✅ UPDATED: Mark as approved
    volunteer.profile.is_approved = True
    volunteer.profile.save()
    
    return Response({
        "message": "Volunteer request accepted successfully",
        "volunteer": VolunteerRequestSerializer(volunteer).data
    })


@api_view(['POST'])
@permission_classes([IsAdmin])
def reject_volunteer_request(request, volunteer_id):
    """
    POST /api/admin/volunteer-requests/{id}/reject/
    Reject a volunteer request
    """
    try:
        volunteer = User.objects.get(id=volunteer_id)
    except User.DoesNotExist:
        return Response(
            {"error": "Volunteer not found"},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Option: Delete the user or mark as rejected
    volunteer_name = volunteer.profile.name
    volunteer.delete()
    
    return Response({
        "message": f"Volunteer request for '{volunteer_name}' rejected successfully"
    })


# ============================================================================
# PERFORMANCE REPORTS
# ============================================================================
@api_view(['GET'])
@permission_classes([IsAdmin])
def projects_progress_report(request):
    """
    GET /api/admin/reports/projects-progress/
    Returns progress report for all projects
    """
    projects = Project.objects.filter(
        status__in=['ACTIVE', 'COMPLETED']
    ).values('title', 'progress')
    
    return Response({
        'projects': [
            {
                'name': p['title'],
                'progress': p['progress']
            }
            for p in projects
        ]
    })


@api_view(['GET'])
@permission_classes([IsAdmin])
def volunteers_performance_report(request):
    """
    GET /api/admin/reports/volunteers-performance/
    Returns performance report for all volunteers
    """
    # ✅ UPDATED: Only show approved volunteers
    volunteers = User.objects.filter(
        profile__role='user',
        profile__is_approved=True
    ).select_related('profile')
    
    data = []
    for volunteer in volunteers:
        completed = volunteer.assigned_tasks.filter(status='مكتملة').count()
        current = volunteer.assigned_tasks.exclude(status='مكتملة').count()
        total = completed + current
        completion_rate = int((completed / total) * 100) if total > 0 else 0
        
        data.append({
            'name': volunteer.profile.name,
            'completed': completed,
            'current': current,
            'completion_rate': completion_rate,
            'join_date': volunteer.date_joined.strftime('%Y-%m-%d'),
        })
    
    return Response({
        'volunteers': data
    })


@api_view(['GET'])
@permission_classes([IsAdmin])
def volunteer_tasks_report(request):
    """
    GET /api/admin/reports/volunteer-tasks/?volunteer_id=1
    Returns tasks for a specific volunteer
    """
    volunteer_id = request.query_params.get('volunteer_id')
    
    if not volunteer_id:
        return Response(
            {"error": "volunteer_id parameter is required"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    tasks = Task.objects.filter(
        volunteer_id=volunteer_id
    ).order_by('-due_date')
    
    serializer = TaskSerializer(tasks, many=True)
    return Response({
        'tasks': serializer.data
    })


# ============================================================================
# SERVICE & SERVICE REQUEST VIEWSETS
# ============================================================================
class ServiceViewSet(viewsets.ModelViewSet):
    """
    Admin endpoint for managing services
    GET /api/services/ - List all services
    POST /api/services/ - Create new service
    """
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer
    permission_classes = [IsAdmin]


class ServiceRequestViewSet(viewsets.ModelViewSet):
    """
    Admin endpoint for managing service requests
    GET /api/service-requests/ - List all service requests
    GET /api/service-requests/?status=PENDING - Filter by status
    """
    queryset = ServiceRequest.objects.all()
    serializer_class = ServiceRequestSerializer
    permission_classes = [IsAdmin]

    def get_queryset(self):
        queryset = super().get_queryset().select_related('service')

        # Filter by status
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)

        return queryset.order_by('-created_at')

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """
        POST /api/service-requests/{id}/approve/
        Approve a service request
        """
        service_request = self.get_object()

        if service_request.status != 'PENDING':
            return Response(
                {"error": "Only pending requests can be approved"},
                status=status.HTTP_400_BAD_REQUEST
            )

        service_request.status = 'APPROVED'
        service_request.save()

        serializer = self.get_serializer(service_request)
        return Response({
            "message": "تم قبول طلب الخدمة بنجاح",
            "request": serializer.data
        })

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """
        POST /api/service-requests/{id}/reject/
        Reject a service request
        """
        service_request = self.get_object()

        if service_request.status != 'PENDING':
            return Response(
                {"error": "Only pending requests can be rejected"},
                status=status.HTTP_400_BAD_REQUEST
            )

        service_request.status = 'REJECTED'
        service_request.save()

        serializer = self.get_serializer(service_request)
        return Response({
            "message": "تم رفض طلب الخدمة",
            "request": serializer.data
        })

    @action(detail=True, methods=['post'])
    def mark_done(self, request, pk=None):
        """
        POST /api/service-requests/{id}/mark_done/
        Mark a service request as done
        """
        service_request = self.get_object()

        if service_request.status not in ['PENDING', 'APPROVED']:
            return Response(
                {"error": "Cannot mark this request as done"},
                status=status.HTTP_400_BAD_REQUEST
            )

        service_request.status = 'DONE'
        service_request.save()

        serializer = self.get_serializer(service_request)
        return Response({
            "message": "تم وضع علامة على الطلب كمكتمل",
            "request": serializer.data
        })


class VolunteerViewSet(viewsets.ModelViewSet):
    queryset = Volunteer.objects.all()
    serializer_class = VolunteerSerializer
    permission_classes = [IsAdmin]


class SuggestionViewSet(viewsets.ModelViewSet):
    queryset = Suggestion.objects.all()
    serializer_class = SuggestionSerializer
    permission_classes = [AllowAny]
    
    def get_permissions(self):
        if self.action in ['update', 'partial_update', 'destroy']:
            return [IsAdmin()]
        return [AllowAny()]


class ProjectAssignmentViewSet(viewsets.ModelViewSet):
    queryset = ProjectAssignment.objects.all()
    serializer_class = ProjectAssignmentSerializer
    permission_classes = [IsAdmin]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        project_id = self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        user_id = self.request.query_params.get('user')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        return queryset


@api_view(['GET'])
@permission_classes([IsAdmin])
def list_users(request):
    """
    GET /api/admin/users/
    Returns list of all registered users with their profiles
    """
    users = User.objects.select_related('profile').all()
    
    data = []
    for user in users:
        data.append({
            'id': user.id,
            'email': user.email,
            'name': user.profile.name,
            'role': user.profile.role,
            'skills': user.profile.skills,
            'city': user.profile.city,
            'total_volunteer_hours': user.profile.total_volunteer_hours,
            'completed_tasks_count': user.profile.completed_tasks_count,
        })
    
    return Response(data)


# ============================================================================
# AUTH VIEWS (Login & Register)
# ============================================================================
class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get("email")
        password = request.data.get("password")

        if not email or not password:
            return Response(
                {"detail": "Email and password are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = authenticate(request, username=email, password=password)

        if user is None:
            return Response(
                {"detail": "بيانات الدخول غير صحيحة"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {
                "name": user.get_full_name() or user.username,
                "email": user.email or user.username,
                "role": "user",
            },
            status=status.HTTP_200_OK,
        )


class RegisterView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        full_name = request.data.get("fullName") or request.data.get("full_name")
        email = request.data.get("email")
        phone = request.data.get("phone")
        password = request.data.get("password")

        if not full_name or not email or not password:
            return Response(
                {"detail": "الاسم، البريد الإلكتروني، وكلمة السر مطلوبة"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if User.objects.filter(username=email).exists():
            return Response(
                {"detail": "هذا البريد مسجّل مسبقاً"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        user = User.objects.create_user(
            username=email,
            email=email,
            password=password,
        )
        user.first_name = full_name
        user.save()

        Volunteer.objects.create(
            full_name=full_name,
            email=email,
            phone=phone or "",
        )

        return Response(
            {
                "name": full_name,
                "email": email,
                "role": "user",
            },
            status=status.HTTP_201_CREATED,
        )


# ============================================================================
# ADMIN REPORTS GENERATION
# ============================================================================
@api_view(['POST'])
@permission_classes([IsAdmin])
def generate_report(request):
    """
    POST /api/admin/reports/generate/
    Generate a comprehensive platform report
    Body: { "date_from": "2026-01-01", "date_to": "2026-01-31" } (optional)
    """
    from datetime import datetime

    date_from = request.data.get('date_from')
    date_to = request.data.get('date_to')

    # Query filters based on date range
    project_filter = Q()
    task_filter = Q()
    volunteer_filter = Q()

    if date_from:
        project_filter &= Q(created_at__gte=date_from)
        task_filter &= Q(created_at__gte=date_from)
        volunteer_filter &= Q(date_joined__gte=date_from)

    if date_to:
        project_filter &= Q(created_at__lte=date_to)
        task_filter &= Q(created_at__lte=date_to)
        volunteer_filter &= Q(date_joined__lte=date_to)

    # ========== COLLECT DATA ==========

    # Projects data
    projects = Project.objects.filter(project_filter)
    projects_by_status = {
        'active': projects.filter(status='ACTIVE').count(),
        'completed': projects.filter(status='COMPLETED').count(),
        'planned': projects.filter(status='PLANNED').count(),
        'cancelled': projects.filter(status='CANCELLED').count(),
    }
    projects_by_category = {}
    for category in projects.values_list('category', flat=True).distinct():
        if category:
            projects_by_category[category] = projects.filter(category=category).count()

    # Projects list with details
    projects_list = []
    for project in projects:
        volunteers_count = ProjectAssignment.objects.filter(project=project).count()
        tasks_count = Task.objects.filter(project=project).count()
        tasks_completed = Task.objects.filter(project=project, status='مكتملة').count()

        # Calculate automatic progress based on task completion
        # If no tasks exist, use manual progress field as fallback
        if tasks_count > 0:
            automatic_progress = int((tasks_completed / tasks_count) * 100)
        else:
            automatic_progress = project.progress  # Fallback to manual progress

        projects_list.append({
            'id': project.id,
            'title': project.title,
            'category': project.category,
            'status': project.status,
            'status_display': {
                'ACTIVE': 'نشط',
                'COMPLETED': 'مكتمل',
                'PLANNED': 'متوقف',
                'CANCELLED': 'ملغي'
            }.get(project.status, project.status),
            'progress': automatic_progress,  # ✅ Now using task-based calculation
            'beneficiaries': project.beneficiaries,
            'donation_amount': float(project.donation_amount),
            'start_date': project.start_date.isoformat() if project.start_date else None,
            'end_date': project.end_date.isoformat() if project.end_date else None,
            'volunteers_assigned': volunteers_count,
            'tasks_total': tasks_count,
            'tasks_completed': tasks_completed,
            'completion_rate': automatic_progress,  # Same value for consistency
        })

    # Volunteers data
    volunteers = User.objects.filter(
        profile__role='user',
        profile__is_approved=True
    ).filter(volunteer_filter).select_related('profile')

    volunteers_list = []
    for volunteer in volunteers:
        tasks = volunteer.assigned_tasks.all()
        tasks_completed = tasks.filter(status='مكتملة').count()
        tasks_in_progress = tasks.exclude(status='مكتملة').count()
        current_projects = list(set([task.project.title for task in tasks.exclude(status='مكتملة')]))

        volunteers_list.append({
            'id': volunteer.id,
            'name': volunteer.profile.name,
            'email': volunteer.email,
            'phone': volunteer.profile.phone,
            'city': volunteer.profile.city,
            'skills': volunteer.profile.skills,
            'qualification': volunteer.profile.qualification,
            'university': volunteer.profile.university,
            'total_hours': volunteer.profile.total_volunteer_hours,
            'tasks_completed': tasks_completed,
            'tasks_in_progress': tasks_in_progress,
            'rating': float(volunteer.profile.rating),
            'join_date': volunteer.date_joined.isoformat(),
            'current_projects': current_projects,
        })

    # Tasks data
    tasks = Task.objects.filter(task_filter).select_related('project', 'volunteer')
    tasks_by_status = {
        'in_progress': tasks.filter(status='قيد التنفيذ').count(),
        'waiting': tasks.filter(status='في الانتظار').count(),
        'completed': tasks.filter(status='مكتملة').count(),
        'on_hold': tasks.filter(status='معلقة').count(),
    }
    tasks_by_priority = {
        'high': tasks.filter(priority='عالية').count(),
        'medium': tasks.filter(priority='متوسطة').count(),
        'low': tasks.filter(priority='منخفضة').count(),
    }

    # Overdue tasks (past due date and not completed)
    from django.utils import timezone
    overdue_tasks = tasks.filter(
        due_date__lt=timezone.now().date(),
        status__in=['قيد التنفيذ', 'في الانتظار', 'معلقة']
    )
    overdue_tasks_list = [{
        'id': task.id,
        'title': task.title,
        'project': task.project.title,
        'volunteer': task.volunteer.profile.name if task.volunteer else None,
        'due_date': task.due_date.isoformat() if task.due_date else None,
        'status': task.status,
        'priority': task.priority,
    } for task in overdue_tasks[:10]]  # Top 10 overdue

    # Totals
    total_beneficiaries = projects.aggregate(total=Sum('beneficiaries'))['total'] or 0
    total_donations = projects.aggregate(total=Sum('donation_amount'))['total'] or 0
    total_volunteer_hours = volunteers.aggregate(
        total=Sum('profile__total_volunteer_hours')
    )['total'] or 0

    # Build report data structure
    report_data = {
        'summary': {
            'total_projects': projects.count(),
            'total_volunteers': volunteers.count(),
            'total_tasks': tasks.count(),
            'total_beneficiaries': total_beneficiaries,
            'total_donations': float(total_donations),
            'total_volunteer_hours': total_volunteer_hours,
            'date_from': date_from,
            'date_to': date_to,
        },
        'projects': {
            'by_status': projects_by_status,
            'by_category': projects_by_category,
            'list': projects_list,
        },
        'volunteers': {
            'total': volunteers.count(),
            'list': volunteers_list,
        },
        'tasks': {
            'by_status': tasks_by_status,
            'by_priority': tasks_by_priority,
            'overdue': overdue_tasks_list,
            'total_completed': tasks_by_status['completed'],
            'completion_rate': int((tasks_by_status['completed'] / tasks.count()) * 100) if tasks.count() > 0 else 0,
        },
    }

    # Generate report title
    now = timezone.now()
    if date_from and date_to:
        title = f"تقرير شامل ({date_from} - {date_to})"
    else:
        title = f"تقرير شامل - {now.strftime('%Y-%m-%d %H:%M')}"

    # Create and save report
    report = AdminReport.objects.create(
        admin=request.user,
        title=title,
        date_from=date_from,
        date_to=date_to,
        report_data=report_data,
        total_projects=projects.count(),
        total_volunteers=volunteers.count(),
        total_tasks=tasks.count(),
        total_beneficiaries=total_beneficiaries,
        total_donations=total_donations,
    )

    serializer = AdminReportSerializer(report)
    return Response({
        'message': 'تم إنشاء التقرير بنجاح',
        'report': serializer.data
    }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAdmin])
def list_reports(request):
    """
    GET /api/admin/reports/
    List all generated reports
    """
    reports = AdminReport.objects.all()
    serializer = AdminReportSerializer(reports, many=True)
    return Response({'results': serializer.data})


@api_view(['GET'])
@permission_classes([IsAdmin])
def get_report_detail(request, report_id):
    """
    GET /api/admin/reports/{id}/
    Get specific report details
    """
    try:
        report = AdminReport.objects.get(id=report_id)
        serializer = AdminReportSerializer(report)
        return Response(serializer.data)
    except AdminReport.DoesNotExist:
        return Response(
            {'error': 'التقرير غير موجود'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['DELETE'])
@permission_classes([IsAdmin])
def delete_report(request, report_id):
    """
    DELETE /api/admin/reports/{id}/
    Delete a report
    """
    try:
        report = AdminReport.objects.get(id=report_id)
        report_title = report.title
        report.delete()
        return Response({
            'message': f'تم حذف التقرير "{report_title}" بنجاح'
        })
    except AdminReport.DoesNotExist:
        return Response(
            {'error': 'التقرير غير موجود'},
            status=status.HTTP_404_NOT_FOUND
        )


# ============================================================================
# VOLUNTEER-SPECIFIC ENDPOINTS (User Pages)
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_volunteer_stats(request):
    """
    GET /api/user/my-stats/
    Returns statistics for the logged-in volunteer
    """
    user = request.user
    profile = user.profile

    # Get completed tasks count
    completed_tasks = user.assigned_tasks.filter(status='مكتملة').count()

    # Get total volunteer hours
    volunteer_hours = profile.total_volunteer_hours or 0

    # Get rating
    rating = float(profile.rating) if profile.rating else 0.0

    # Calculate points (simple formula: hours + completed_tasks * 10)
    points = int(volunteer_hours) + (completed_tasks * 10)

    return Response({
        'volunteer_hours': volunteer_hours,
        'rating': rating,
        'completed_tasks': completed_tasks,
        'points': points,
        'name': profile.name,
        'email': user.email
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def my_tasks(request):
    """
    GET /api/user/my-tasks/
    Returns tasks assigned to the logged-in volunteer
    """
    user = request.user

    # Get tasks assigned to this volunteer, excluding completed ones for the main page
    tasks = user.assigned_tasks.select_related('project').prefetch_related('subtasks').all()

    serializer = TaskSerializer(tasks, many=True)
    return Response({
        'results': serializer.data
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def public_projects(request):
    """
    GET /api/public-projects/
    Public endpoint - returns all visible projects for public viewing
    No authentication required
    """
    # Get all visible projects (not hidden)
    projects = Project.objects.filter(
        is_hidden=False
    ).order_by('-created_at')

    serializer = ProjectSerializer(projects, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def available_opportunities(request):
    """
    GET /api/user/opportunities/
    Returns available projects/opportunities that volunteers can apply to
    Shows all projects that are NOT hidden (is_hidden=False)
    Volunteers can apply and wait for admin approval
    Requires authentication
    """
    # Get all visible projects (regardless of status)
    opportunities = Project.objects.filter(
        is_hidden=False
    ).order_by('-created_at')  # Show all visible projects, sorted by newest first

    serializer = ProjectSerializer(opportunities, many=True)
    return Response({
        'results': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def apply_to_opportunity(request, project_id):
    """
    POST /api/user/opportunities/{project_id}/apply/
    Apply to an opportunity/project - Creates an application for admin review
    """
    try:
        project = Project.objects.get(id=project_id)
        user = request.user

        # Check if already applied
        existing_application = VolunteerApplication.objects.filter(
            project=project,
            volunteer=user
        ).first()

        if existing_application:
            return Response(
                {'message': 'لقد تقدمت لهذه الفرصة من قبل'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create a new application (NOT a task yet)
        application = VolunteerApplication.objects.create(
            volunteer=user,
            project=project,
            message=request.data.get('message', ''),
            status='قيد المراجعة'
        )

        return Response({
            'message': 'تم إرسال طلبك بنجاح. سيتم مراجعته من قبل المشرف.',
            'application_id': application.id
        })

    except Project.DoesNotExist:
        return Response(
            {'error': 'المشروع غير موجود'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def withdraw_from_task(request, task_id):
    """
    POST /api/user/tasks/{task_id}/withdraw/
    Withdraw from a task - marks it as cancelled instead of deleting
    """
    try:
        task = Task.objects.get(id=task_id, volunteer=request.user)

        # Don't delete - just mark as cancelled so it appears in cancelled tab
        task.status = 'ملغاة'
        task.save()

        return Response({
            'message': f'تم الانسحاب من المهمة "{task.title}" بنجاح'
        })

    except Task.DoesNotExist:
        return Response(
            {'error': 'المهمة غير موجودة أو ليس لديك صلاحية للانسحاب منها'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_task_progress(request, task_id):
    """
    PATCH /api/user/tasks/{task_id}/update-progress/
    Update task progress and subtasks
    Body: {
        "progress": 50,
        "subtasks": [
            {"text": "Task 1", "done": true},
            {"text": "Task 2", "done": false}
        ]
    }
    """
    try:
        task = Task.objects.get(id=task_id, volunteer=request.user)

        # Track if task was already completed before this update
        was_completed = task.status == 'مكتملة'

        # Update progress if provided
        if 'progress' in request.data:
            task.progress = request.data['progress']

            # Auto-update status based on progress
            if task.progress == 100:
                task.status = 'مكتملة'
            elif task.progress > 0:
                task.status = 'قيد التنفيذ'

        # Update subtasks if provided
        if 'subtasks' in request.data:
            # Delete existing subtasks
            task.subtasks.all().delete()

            # Create new subtasks
            for idx, subtask_data in enumerate(request.data['subtasks']):
                Subtask.objects.create(
                    task=task,
                    title=subtask_data.get('text', ''),
                    completed=subtask_data.get('done', False),
                    order=idx
                )

        task.save()

        # Update volunteer profile stats based on completion status change
        is_now_completed = task.status == 'مكتملة'
        profile = request.user.profile

        # Task just became completed - add stats
        if is_now_completed and not was_completed:
            profile.total_volunteer_hours += task.hours or 0
            profile.completed_tasks_count += 1
            profile.points = profile.total_volunteer_hours + (profile.completed_tasks_count * 10)
            profile.save()

        # Task was un-completed - subtract stats
        elif was_completed and not is_now_completed:
            profile.total_volunteer_hours = max(0, profile.total_volunteer_hours - (task.hours or 0))
            profile.completed_tasks_count = max(0, profile.completed_tasks_count - 1)
            profile.points = profile.total_volunteer_hours + (profile.completed_tasks_count * 10)
            profile.save()

        serializer = TaskSerializer(task)
        return Response({
            'message': 'تم تحديث المهمة بنجاح',
            'task': serializer.data
        })

    except Task.DoesNotExist:
        return Response(
            {'error': 'المهمة غير موجودة أو ليس لديك صلاحية لتحديثها'},
            status=status.HTTP_404_NOT_FOUND
        )


# ============================================================================
# VOLUNTEER APPLICATION MANAGEMENT (ADMIN)
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAdmin])
def list_volunteer_applications(request):
    """
    GET /api/admin/applications/
    List all volunteer applications with optional status filter
    """
    status_filter = request.query_params.get('status', None)

    applications = VolunteerApplication.objects.select_related(
        'volunteer__profile', 'project', 'reviewed_by__profile'
    ).all()

    if status_filter:
        applications = applications.filter(status=status_filter)

    serializer = VolunteerApplicationSerializer(applications, many=True)
    return Response({'results': serializer.data})


@api_view(['POST'])
@permission_classes([IsAdmin])
def accept_volunteer_application(request, application_id):
    """
    POST /api/admin/applications/{application_id}/accept/
    Accept a volunteer application and create a task for the volunteer
    """
    try:
        application = VolunteerApplication.objects.get(id=application_id)

        if application.status != 'قيد المراجعة':
            return Response(
                {'error': 'هذا الطلب تم مراجعته بالفعل'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update application status
        application.status = 'مقبول'
        application.reviewed_by = request.user
        application.reviewed_at = timezone.now()
        application.admin_notes = request.data.get('admin_notes', '')
        application.save()

        # Create a task for the volunteer
        task = Task.objects.create(
            title=f"مهمة في {application.project.title}",
            description=application.project.desc or "لا يوجد وصف",
            project=application.project,
            volunteer=application.volunteer,
            status='قيد التنفيذ',  # Accepted = In Progress
            priority='متوسطة',
            hours=application.project.estimated_hours or 0
        )

        # Create default subtasks
        default_subtasks = [
            "مراجعة متطلبات المهمة",
            "البدء في التنفيذ",
            "إتمام المهمة ومراجعتها",
        ]

        for idx, subtask_title in enumerate(default_subtasks):
            Subtask.objects.create(
                task=task,
                title=subtask_title,
                completed=False,
                order=idx
            )

        return Response({
            'message': 'تم قبول الطلب وإنشاء مهمة للمتطوع',
            'task_id': task.id
        })

    except VolunteerApplication.DoesNotExist:
        return Response(
            {'error': 'الطلب غير موجود'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['POST'])
@permission_classes([IsAdmin])
def reject_volunteer_application(request, application_id):
    """
    POST /api/admin/applications/{application_id}/reject/
    Reject a volunteer application
    """
    try:
        application = VolunteerApplication.objects.get(id=application_id)

        if application.status != 'قيد المراجعة':
            return Response(
                {'error': 'هذا الطلب تم مراجعته بالفعل'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update application status
        application.status = 'مرفوض'
        application.reviewed_by = request.user
        application.reviewed_at = timezone.now()
        application.admin_notes = request.data.get('admin_notes', '')
        application.save()

        return Response({
            'message': 'تم رفض الطلب'
        })

    except VolunteerApplication.DoesNotExist:
        return Response(
            {'error': 'الطلب غير موجود'},
            status=status.HTTP_404_NOT_FOUND
        )

# ============================================================================
# PUBLIC ENDPOINTS (No Authentication Required)
# ============================================================================

@api_view(['GET'])
@permission_classes([AllowAny])
def public_volunteers_stats(request):
    """
    GET /api/public-volunteers-stats/
    Returns volunteer statistics for the public volunteers page.
    No authentication required.
    """
    from accounts.models import Profile

    # Get all user profiles (volunteers only)
    volunteers = Profile.objects.filter(role='user')

    stats = []
    for profile in volunteers:
        # Calculate stats for each volunteer
        total_hours = Task.objects.filter(
            volunteer=profile.user,
            status='مكتملة'
        ).aggregate(total=Sum('hours'))['total'] or 0

        participations_count = Task.objects.filter(
            volunteer=profile.user
        ).exclude(status='ملغاة').count()

        successes_count = Task.objects.filter(
            volunteer=profile.user,
            status='مكتملة'
        ).count()

        stats.append({
            'id': profile.user.id,
            'gender': profile.gender,
            'total_hours': total_hours,
            'participations_count': participations_count,
            'successes_count': successes_count,
        })

    return Response(stats)


@api_view(['POST'])
@permission_classes([AllowAny])
def public_submit_suggestion(request):
    """
    POST /api/public-suggestions/
    Submit a suggestion from the public suggest page
    No authentication required

    Payload:
    {
        "title": "Suggestion title",
        "description": "Suggestion description",
        "submitted_by": "email@example.com"
    }
    """
    serializer = SuggestionSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response({
            'message': 'تم استلام اقتراحك بنجاح'
        }, status=status.HTTP_201_CREATED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([AllowAny])
def public_home_stats(request):
    """
    GET /api/public-home-stats/
    Returns aggregated statistics for home page (Hero section)
    No authentication required
    """
    # Total beneficiaries from all projects
    total_beneficiaries = Project.objects.aggregate(
        total=Sum('beneficiaries')
    )['total'] or 0

    # Total potential projects (all non-cancelled projects)
    potential_projects = Project.objects.exclude(status='CANCELLED').count()

    # Total donations amount
    total_donations = Project.objects.aggregate(
        total=Sum('donation_amount')
    )['total'] or 0

    return Response({
        'beneficiaries': total_beneficiaries,
        'potential_projects': potential_projects,
        'donations': float(total_donations),
    })


@api_view(['POST'])
@permission_classes([AllowAny])
def public_submit_service_request(request):
    """
    POST /api/public-service-request/
    Submit a service request from the public (beneficiaries)
    No authentication required

    Payload:
    {
        "service": 1,  // Service ID
        "beneficiary_name": "اسم المسجد",
        "beneficiary_contact": "0500000000",
        "details": "نحتاج إلى 100 زجاجة ماء للمسجد"
    }
    """
    serializer = ServiceRequestSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
        return Response({
            'message': 'تم استلام طلبك بنجاح. سيتم مراجعته من قبل الإدارة.',
            'request_id': serializer.data['id']
        }, status=status.HTTP_201_CREATED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([AllowAny])
def public_services_list(request):
    """
    GET /api/public-services/
    List all active VOLUNTEER OPPORTUNITY services for /services page
    No authentication required
    """
    services = Service.objects.filter(
        is_active=True,
        service_type="للمتطوعين"  # Only volunteer opportunity services
    ).order_by('-created_at')
    serializer = ServiceSerializer(services, many=True)
    return Response({
        'results': serializer.data
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def beneficiary_services_list(request):
    """
    GET /api/beneficiary-services/
    List all active BENEFICIARY services for main page
    No authentication required
    """
    services = Service.objects.filter(
        is_active=True,
        service_type="للمستفيدين"  # Only beneficiary services
    ).order_by('-created_at')
    serializer = ServiceSerializer(services, many=True)
    return Response({
        'results': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def apply_to_service_as_volunteer(request, service_id):
    """
    POST /api/services/{service_id}/apply-volunteer/
    Apply to help with a service as a volunteer
    Requires authentication
    """
    try:
        service = Service.objects.get(id=service_id)
        user = request.user

        # Check if already applied
        existing_application = ServiceVolunteerApplication.objects.filter(
            service=service,
            volunteer=user
        ).first()

        if existing_application:
            return Response(
                {'message': 'لقد تقدمت للمساعدة في هذه الخدمة من قبل'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create new application
        application = ServiceVolunteerApplication.objects.create(
            volunteer=user,
            service=service,
            message=request.data.get('message', ''),
            status='قيد المراجعة'
        )

        return Response({
            'message': 'تم إرسال طلبك للمساعدة بنجاح. سيتم مراجعته من قبل المشرف.',
            'application_id': application.id
        })

    except Service.DoesNotExist:
        return Response(
            {'error': 'الخدمة غير موجودة'},
            status=status.HTTP_404_NOT_FOUND
        )


# ============================================================================
# ADMIN: SERVICE VOLUNTEER APPLICATIONS MANAGEMENT
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAdmin])
def list_service_volunteer_applications(request):
    """
    GET /api/admin/service-volunteer-applications/
    List all volunteer applications for services with optional status filter
    """
    status_filter = request.query_params.get('status', None)

    applications = ServiceVolunteerApplication.objects.select_related(
        'volunteer__profile', 'service', 'reviewed_by__profile'
    ).all()

    if status_filter:
        applications = applications.filter(status=status_filter)

    serializer = ServiceVolunteerApplicationSerializer(applications, many=True)
    return Response({'results': serializer.data})


@api_view(['POST'])
@permission_classes([IsAdmin])
def accept_service_volunteer_application(request, application_id):
    """
    POST /api/admin/service-volunteer-applications/{application_id}/accept/
    Accept a volunteer application for a service
    """
    try:
        application = ServiceVolunteerApplication.objects.get(id=application_id)

        if application.status != 'قيد المراجعة':
            return Response(
                {'error': 'هذا الطلب تم مراجعته بالفعل'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update application status
        application.status = 'مقبول'
        application.reviewed_by = request.user
        application.reviewed_at = timezone.now()
        application.admin_notes = request.data.get('admin_notes', '')
        application.save()

        return Response({
            'message': 'تم قبول المتطوع للمساعدة في الخدمة'
        })

    except ServiceVolunteerApplication.DoesNotExist:
        return Response(
            {'error': 'الطلب غير موجود'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['POST'])
@permission_classes([IsAdmin])
def reject_service_volunteer_application(request, application_id):
    """
    POST /api/admin/service-volunteer-applications/{application_id}/reject/
    Reject a volunteer application for a service
    """
    try:
        application = ServiceVolunteerApplication.objects.get(id=application_id)

        if application.status != 'قيد المراجعة':
            return Response(
                {'error': 'هذا الطلب تم مراجعته بالفعل'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update application status
        application.status = 'مرفوض'
        application.reviewed_by = request.user
        application.reviewed_at = timezone.now()
        application.admin_notes = request.data.get('admin_notes', '')
        application.save()

        return Response({
            'message': 'تم رفض الطلب'
        })

    except ServiceVolunteerApplication.DoesNotExist:
        return Response(
            {'error': 'الطلب غير موجود'},
            status=status.HTTP_404_NOT_FOUND
        )


# ============================================================================
# PUBLIC: VOLUNTEER STATISTICS ENDPOINT
# ============================================================================

@api_view(['GET'])
@permission_classes([AllowAny])
def public_volunteer_statistics(request):
    """
    GET /api/public-volunteer-statistics/
    Returns volunteer statistics for the home page dashboard
    Optional query param: ?year=2025 (defaults to latest year)
    """
    year = request.query_params.get('year', None)

    if year:
        stats = VolunteerStatistics.objects.filter(year=int(year)).first()
    else:
        # Get the most recent year's statistics
        stats = VolunteerStatistics.objects.first()

    if not stats:
        return Response({
            'error': 'No statistics found',
            'data': None
        }, status=status.HTTP_404_NOT_FOUND)

    serializer = VolunteerStatisticsSerializer(stats)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([AllowAny])
def public_water_supply_request(request):
    """
    POST /api/public-water-supply-request/
    Submit a water supply request for a mosque
    No authentication required
    """
    data = request.data.copy()

    # Convert frontend field names to backend field names
    field_mapping = {
        'applicantName': 'applicant_name',
        'mobileNumber': 'mobile_number',
        'applicantRole': 'applicant_role',
        'mosqueName': 'mosque_name',
        'neighborhood': 'neighborhood',
        'locationLink': 'location_link',
        'worshippersCount': 'worshippers_count',
        'donorExists': 'donor_exists',
        'donorName': 'donor_name',
        'donorPhone': 'donor_phone',
    }

    # Map frontend camelCase to backend snake_case
    mapped_data = {}
    for frontend_key, backend_key in field_mapping.items():
        if frontend_key in data:
            value = data[frontend_key]
            # Convert donorExists from string to boolean
            if frontend_key == 'donorExists':
                value = value == 'نعم'
            # Convert worshippersCount to integer
            elif frontend_key == 'worshippersCount':
                try:
                    value = int(value)
                except (ValueError, TypeError):
                    value = 0
            mapped_data[backend_key] = value

    serializer = WaterSupplyRequestSerializer(data=mapped_data)
    if serializer.is_valid():
        serializer.save()
        return Response({
            'success': True,
            'message': 'تم إرسال طلبك بنجاح'
        }, status=status.HTTP_201_CREATED)

    return Response({
        'success': False,
        'errors': serializer.errors
    }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAdmin])
def upload_volunteer_statistics(request):
    """
    POST /api/admin/upload-statistics/
    Upload Excel file to update volunteer statistics for home page
    Requires admin authentication
    """
    if 'file' not in request.FILES:
        return Response({
            'success': False,
            'error': 'لم يتم رفع أي ملف'
        }, status=status.HTTP_400_BAD_REQUEST)

    excel_file = request.FILES['file']

    try:
        # Read Excel file
        wb = openpyxl.load_workbook(BytesIO(excel_file.read()), data_only=True)
        ws = wb.active

        # Get all rows (skip header)
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        # Calculate statistics from raw data
        total_records = len([r for r in rows if r and any(r)])

        # Count unique volunteers by phone or ID
        volunteer_ids = set()
        total_hours = 0
        department_hours_map = {}
        volunteer_hours_map = {}

        # Find column indices from header
        headers = [cell.value for cell in ws[1]]

        # Try to find relevant columns
        hours_col = None
        dept_col = None
        name_col = None
        id_col = None

        for i, h in enumerate(headers):
            if h:
                h_lower = str(h).strip()
                if 'ساع' in h_lower or 'hours' in h_lower.lower():
                    hours_col = i
                elif 'إدار' in h_lower or 'قسم' in h_lower or 'department' in h_lower.lower():
                    dept_col = i
                elif 'اسم' in h_lower and 'مشروع' not in h_lower:
                    name_col = i
                elif 'هوية' in h_lower or 'id' in h_lower.lower():
                    id_col = i

        for row in rows:
            if not row or not any(row):
                continue

            # Get volunteer identifier
            vol_id = None
            if id_col is not None and id_col < len(row):
                vol_id = row[id_col]
            if vol_id:
                volunteer_ids.add(str(vol_id))

            # Get hours
            hours = 0
            if hours_col is not None and hours_col < len(row):
                try:
                    hours = float(row[hours_col] or 0)
                except (ValueError, TypeError):
                    hours = 0
            total_hours += hours

            # Get department
            dept = 'غير مسند'
            if dept_col is not None and dept_col < len(row):
                dept = str(row[dept_col] or 'غير مسند').strip()

            if dept not in department_hours_map:
                department_hours_map[dept] = 0
            department_hours_map[dept] += hours

            # Track volunteer hours for top volunteers
            vol_name = None
            if name_col is not None and name_col < len(row):
                vol_name = str(row[name_col] or '').strip()
            if vol_name:
                if vol_name not in volunteer_hours_map:
                    volunteer_hours_map[vol_name] = 0
                volunteer_hours_map[vol_name] += hours

        # Get or create 2025 statistics
        year = 2025
        stats, created = VolunteerStatistics.objects.update_or_create(
            year=year,
            defaults={
                'total_volunteers': len(volunteer_ids) or total_records,
                'new_volunteers': int(len(volunteer_ids) * 0.78),  # Estimate 78% new
                'returning_volunteers': int(len(volunteer_ids) * 0.22),  # Estimate 22% returning
                'total_hours': int(total_hours),
                'total_contribution_value': Decimal(total_hours * 13),  # 13 SAR per hour
                'contribution_value_display': f"{total_hours * 13 / 1000000:.2f}M" if total_hours > 100000 else f"{total_hours * 13 / 1000:.0f}K",
            }
        )

        # Clear and recreate department hours
        DepartmentHours.objects.filter(statistics=stats).delete()

        colors = ['#6B1F2B', '#8B5A2B', '#2E8B57', '#4169E1', '#9370DB', '#FF8C00', '#20B2AA', '#DC143C']
        total_dept_hours = sum(department_hours_map.values()) or 1

        for i, (dept, hours) in enumerate(sorted(department_hours_map.items(), key=lambda x: -x[1])):
            DepartmentHours.objects.create(
                statistics=stats,
                department_name=dept,
                department_name_ar=dept,
                hours=int(hours),
                percentage=Decimal(hours / total_dept_hours * 100),
                color=colors[i % len(colors)]
            )

        # Clear and recreate top volunteers
        TopVolunteer.objects.filter(statistics=stats).delete()

        top_vols = sorted(volunteer_hours_map.items(), key=lambda x: -x[1])[:5]
        for rank, (name, hours) in enumerate(top_vols, 1):
            TopVolunteer.objects.create(
                statistics=stats,
                rank=rank,
                name=name,
                hours=int(hours)
            )

        return Response({
            'success': True,
            'message': 'تم تحديث الإحصائيات بنجاح',
            'data': {
                'total_records': total_records,
                'total_volunteers': len(volunteer_ids) or total_records,
                'total_hours': int(total_hours),
                'departments': len(department_hours_map),
                'top_volunteers': len(top_vols)
            }
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({
            'success': False,
            'error': f'خطأ في معالجة الملف: {str(e)}'
        }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT'])
@permission_classes([IsAdmin])
def admin_volunteer_statistics(request):
    """
    GET/PUT /api/admin/volunteer-statistics/
    Get or update volunteer statistics for home page dashboard
    Requires admin authentication
    """
    year = request.query_params.get('year', 2025)

    if request.method == 'GET':
        stats = VolunteerStatistics.objects.filter(year=year).first()
        if not stats:
            return Response({'error': 'No statistics found'}, status=status.HTTP_404_NOT_FOUND)
        serializer = VolunteerStatisticsSerializer(stats)
        return Response(serializer.data)

    elif request.method == 'PUT':
        data = request.data

        # Get or create statistics for the year
        stats, created = VolunteerStatistics.objects.update_or_create(
            year=int(data.get('year', 2025)),
            defaults={
                'total_volunteers': int(data.get('total_volunteers', 0)),
                'new_volunteers': int(data.get('new_volunteers', 0)),
                'returning_volunteers': int(data.get('returning_volunteers', 0)),
                'total_hours': int(data.get('total_hours', 0)),
                'total_contribution_value': Decimal(str(data.get('total_contribution_value', 0))),
                'contribution_value_display': data.get('contribution_value_display', ''),
            }
        )

        # Update quarterly targets if provided
        quarterly_targets = data.get('quarterly_targets', [])
        if quarterly_targets:
            QuarterlyTarget.objects.filter(statistics=stats).delete()
            for qt in quarterly_targets:
                QuarterlyTarget.objects.create(
                    statistics=stats,
                    quarter=int(qt.get('quarter', 1)),
                    volunteer_target=int(qt.get('volunteer_target', 0)),
                    volunteer_actual=int(qt.get('volunteer_actual', 0)),
                    hours_target=int(qt.get('hours_target', 0)),
                    hours_actual=int(qt.get('hours_actual', 0)),
                )

        # Update department hours if provided
        department_hours = data.get('department_hours', [])
        if department_hours:
            DepartmentHours.objects.filter(statistics=stats).delete()
            for dh in department_hours:
                DepartmentHours.objects.create(
                    statistics=stats,
                    department_name=dh.get('department_name', ''),
                    department_name_ar=dh.get('department_name_ar', dh.get('label', '')),
                    hours=int(dh.get('hours', dh.get('value', 0))),
                    percentage=Decimal(str(dh.get('percentage', 0))),
                    color=dh.get('color', '#6B1F2B'),
                )

        # Update top volunteers if provided
        top_volunteers = data.get('top_volunteers', [])
        if top_volunteers:
            TopVolunteer.objects.filter(statistics=stats).delete()
            for tv in top_volunteers:
                TopVolunteer.objects.create(
                    statistics=stats,
                    rank=int(tv.get('rank', 1)),
                    name=tv.get('name', ''),
                    hours=int(tv.get('hours', 0)),
                )

        serializer = VolunteerStatisticsSerializer(stats)
        return Response({
            'success': True,
            'message': 'تم تحديث الإحصائيات بنجاح',
            'data': serializer.data
        }, status=status.HTTP_200_OK)
